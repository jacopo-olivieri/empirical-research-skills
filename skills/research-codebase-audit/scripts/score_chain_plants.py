#!/usr/bin/env python3
"""Campaign-close scorer for the U18 chain plants.

One plant per invocation: reads ``<plant>/expected.json`` (the answer key)
and a finished plant run's ``audit/`` directory, and applies three
fail-closed gate legs:

1. **Presence + pass condition:** the final code-error register contains at
   least one row matching the plant's mechanism signature whose ``Status``
   is in the key's gate-counting terminal set AND whose ``Severity`` meets
   the key's floor. A matching row at any other status is reported ("row
   present at <status>") and is RED — mere presence does not pass. A
   ``duplicate_of:<ID>`` tombstone is followed to its target row, which
   passes only if the target ALSO matches the plant signature and meets the
   status and severity bar (a tombstone onto an unrelated confirmed row
   must not satisfy the gate).
2. **Workbook shape:** ``audit/code_review.xlsx`` exists; its sheet set
   equals the expected set exactly; the Paper Claims header row equals the
   pinned visible list in order and contains none of the hidden names; the
   Code Errors header row equals the pinned list.
3. **Effort map:** ``audit/_run/manifest.json``'s ``effort_map`` equals the
   key's pinned literal exactly.

Every answer-key field is required: the key schema is validated up front
and any missing or malformed block is exit 2 — an incomplete key must
never silently disable a gate leg. A missing, unparseable, or malformed
audit input (register, workbook, manifest) is RED, never GREEN.

Register parsing and signature matching reuse ``score_fixture.py``'s
helpers (``load_rows``, ``row_text``, ``sig_match``, ``severity``); the
qualifying-status logic is deliberately NOT reused — the chain-plant pass
condition is stricter than the fixture scorer's ``qualifies``.

On a RED, the operator diagnoses the broken chain link by hand from the
named production artifacts (emitter output, mapping, recheck ledger);
there is no automated trace.

Usage:
    score_chain_plants.py --plant fixture/chain_plants/<name> --audit-dir PATH

Exit codes: 0 = PLANT GREEN, 1 = PLANT RED, 2 = usage/IO error.
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from score_fixture import load_rows, row_text, severity, sig_match


class KeyError_(ValueError):
    """The answer key is missing, unreadable, or fails schema validation."""


# --------------------------------------------------------------- answer key


def _require(condition, message):
    if not condition:
        raise KeyError_(message)


def _is_str_list(value):
    return (isinstance(value, list) and value
            and all(isinstance(item, str) and item for item in value))


WORKBOOK_KEYS = ("sheets", "paper_claims_visible", "paper_claims_hidden_absent",
                 "code_errors_headers")


def load_key(plant_dir):
    """Load and schema-validate ``<plant>/expected.json``.

    Every field is required; a missing or malformed block raises (exit 2 in
    ``main``) rather than silently disabling the gate leg it feeds.
    """
    path = Path(plant_dir) / "expected.json"
    _require(path.is_file(), f"answer key missing: {path}")
    try:
        key = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise KeyError_(f"answer key invalid JSON: {exc}")
    _require(isinstance(key, dict), "answer key must be a JSON object")
    for field in ("class", "run8_id"):
        _require(isinstance(key.get(field), str) and key[field],
                 f"answer key field '{field}' must be a non-empty string")
    signature = key.get("signature")
    _require(isinstance(signature, list) and signature
             and all(_is_str_list(group) for group in signature),
             "answer key field 'signature' must be a non-empty list of "
             "non-empty string groups")
    _require(_is_str_list(key.get("statuses")),
             "answer key field 'statuses' must be a non-empty string list")
    _require(isinstance(key.get("min_severity"), int)
             and not isinstance(key.get("min_severity"), bool),
             "answer key field 'min_severity' must be an integer")
    workbook = key.get("workbook")
    _require(isinstance(workbook, dict),
             "answer key field 'workbook' must be an object")
    for field in WORKBOOK_KEYS:
        _require(_is_str_list(workbook.get(field)),
                 f"answer key workbook field '{field}' must be a non-empty "
                 "string list")
    effort = key.get("effort_map")
    _require(isinstance(effort, dict) and effort
             and all(isinstance(k, str) and isinstance(v, str) and k and v
                     for k, v in effort.items()),
             "answer key field 'effort_map' must be a non-empty object of "
             "string pairs")
    return key


# ---------------------------------------------------------------- gate legs


def _passes(row, key):
    """True when *row* sits at a gate-counting status and severity."""
    status = (row.get("Status") or "").strip()
    return status in key["statuses"] and severity(row) >= key["min_severity"]


def check_register(audit, key, reasons):
    """Leg 1: presence + pass condition on the final code-error register."""
    path = audit / "code_error_register.md"
    if not path.is_file():
        reasons.append("code_error_register.md missing")
        return
    rows = load_rows(path, "Error ID")
    if not rows:
        reasons.append("no code-error table parsed from "
                       "code_error_register.md")
        return
    matches = [row for row in rows
               if sig_match(row_text(row), key["signature"])]
    if not matches:
        reasons.append("no register row matches the plant signature")
        return
    if any(_passes(row, key) for row in matches):
        return
    # Follow duplicate_of tombstones one hop: the register contract permits
    # duplicates only for same mechanism and location, so a passing
    # equivalent target matches the signature by construction; a target
    # that does not match the signature is an unrelated row and fails.
    by_id = {(row.get("Error ID") or "").strip(): row for row in rows}
    for row in matches:
        status = (row.get("Status") or "").strip()
        if not status.startswith("duplicate_of:"):
            continue
        target = by_id.get(status.split(":", 1)[1].strip())
        if target is None:
            continue  # dangling tombstone: fail-closed via the report below
        if sig_match(row_text(target), key["signature"]) and _passes(target, key):
            return
    found = "; ".join(
        f"{(row.get('Error ID') or '?').strip()} "
        f"status={(row.get('Status') or '?').strip() or '?'} "
        f"sev={(row.get('Severity') or '-').strip() or '-'}"
        for row in matches)
    reasons.append(
        f"row present at {found} — not at a gate-counting status "
        f"({'/'.join(key['statuses'])}) with severity >= {key['min_severity']}")


def _header_row(worksheet):
    row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    cells = list(row)
    while cells and cells[-1] is None:
        cells.pop()
    return [str(cell) for cell in cells]


def check_workbook(audit, key, reasons):
    """Leg 2: exported workbook shape (sheets + pinned header rows)."""
    expected = key["workbook"]
    path = audit / "code_review.xlsx"
    if not path.is_file():
        reasons.append("code_review.xlsx missing")
        return
    try:
        workbook = load_workbook(path, read_only=True)
    except Exception as exc:  # malformed/unreadable file: RED, never GREEN
        reasons.append(f"code_review.xlsx unreadable: {exc}")
        return
    try:
        if set(workbook.sheetnames) != set(expected["sheets"]):
            reasons.append(
                f"sheet set {sorted(workbook.sheetnames)} != expected "
                f"{sorted(expected['sheets'])}")
        if "Paper Claims" in workbook.sheetnames:
            got = _header_row(workbook["Paper Claims"])
            if got != expected["paper_claims_visible"]:
                reasons.append(
                    f"Paper Claims header {got} != pinned visible list "
                    f"{expected['paper_claims_visible']}")
            hidden = [name for name in expected["paper_claims_hidden_absent"]
                      if name in got]
            if hidden:
                reasons.append(
                    f"hidden column(s) present on Paper Claims: {hidden}")
        else:
            reasons.append("Paper Claims sheet missing")
        if "Code Errors" in workbook.sheetnames:
            got = _header_row(workbook["Code Errors"])
            if got != expected["code_errors_headers"]:
                reasons.append(
                    f"Code Errors header {got} != pinned list "
                    f"{expected['code_errors_headers']}")
        else:
            reasons.append("Code Errors sheet missing")
    finally:
        workbook.close()


def check_effort_map(audit, key, reasons):
    """Leg 3: the run manifest records the pinned effort map exactly."""
    path = audit / "_run" / "manifest.json"
    if not path.is_file():
        reasons.append("_run/manifest.json missing")
        return
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        reasons.append(f"manifest.json invalid JSON: {exc}")
        return
    if not isinstance(manifest, dict):
        reasons.append("manifest.json is not a JSON object")
        return
    if manifest.get("effort_map") != key["effort_map"]:
        reasons.append(
            "manifest effort_map differs from the answer key's pinned map")


# --------------------------------------------------------------------- main


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--plant", type=Path, required=True,
                        help="plant directory holding expected.json")
    parser.add_argument("--audit-dir", type=Path, required=True,
                        help="the finished plant run's audit/ directory")
    args = parser.parse_args()

    try:
        key = load_key(args.plant)
    except KeyError_ as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    if not args.audit_dir.is_dir():
        print(f"ERROR: audit dir missing: {args.audit_dir}", file=sys.stderr)
        return 2

    reasons = []
    check_register(args.audit_dir, key, reasons)
    check_workbook(args.audit_dir, key, reasons)
    check_effort_map(args.audit_dir, key, reasons)

    if reasons:
        print(f"PLANT RED — {key['class']}: " + " | ".join(reasons))
        return 1
    print(f"PLANT GREEN — {key['class']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
