"""U18 — chain-plant scorer tests (marker ``u18``).

The scorer's gate is Tier-1: its silent failure would let the campaign
claim closure on plants that did not actually pass. The suite proves, per
gate leg:

- **Test 2 (good input):** for each of the three plants, a synthetic
  completed plant-run ``audit/`` dir — final registers carrying the plant
  row, a workbook produced by the REAL ``export_xlsx.py``, and a manifest
  carrying the default effort map — scores GREEN through the production
  CLI.
- **Test 1 (bad input):** every sabotaged or degraded input is RED (or
  exit 2 for a broken answer key), each with the right reason.
- **Test 3 (leg is load-bearing):** with a schema-valid answer key that
  weakens one leg, the corresponding Test-1 sabotage scores GREEN —
  proving Test 1's RED comes from the leg, not from an accident.
- **Sabotage drills through the production CLI, one per gate leg:**
  status-flip, severity-drop, and row-deletion (hand-edited register), a
  stray workbook sheet, and a flipped manifest effort value.
- **Lockstep guard:** each answer key's ``effort_map`` literal equals
  ``dispatch_tracking.DEFAULT_EFFORT_MAP`` by value, so the keys and a
  future default change cannot drift apart silently.

The synthetic run dirs are built once per plant (module scope) and copied
per test; all fixture content here is fresh synthetic material in the
three plant domains (school-meals nutrition, air-quality exceedance,
interlibrary loans).
"""

import json
import shutil

import pytest

import regbuild as rb

pytestmark = pytest.mark.u18

SCORER = "score_chain_plants.py"
CHAIN_PLANTS = rb.FIXTURE_DIR / "chain_plants"
PLANTS = ("comment_closure", "producer_group", "path_derivation")

# One register row per plant whose text satisfies the plant's mechanism
# signature at the gate-counting terminal state (confirmed, severity 2).
PLANT_ROWS = {
    "comment_closure": rb.error_row(
        "E-0201", etype="sample_filter_or_flag_error",
        source="`do/build_sample.do`", location="`do/build_sample.do:11`",
        desc=("the keep guard adds a meal_plan == standard conjunct that "
              "narrows the cases eligible_flag's own comment says are "
              "covered, silently dropping every reduced-price student"),
        why=("the narrowed sample flows through output/analysis_sample.dta "
             "into the Table 1 mean protein intake")),
    "producer_group": rb.error_row(
        "E-0202", etype="sample_filter_or_flag_error",
        source="`do/build_flags.do`", location="`do/build_flags.do:13-16`",
        desc=("each foreach iteration overwrites exceed_any by plain "
              "assignment, so the final pollutant's write erases every "
              "earlier exceedance the first pollutant set"),
        why=("the understated share of exceedance station-days flows into "
             "the reported Table 1 number")),
    "path_derivation": rb.error_row(
        "E-0203", etype="stale_or_wrong_path",
        source="`py/tools/make_totals.py`",
        location="`py/tools/make_totals.py:10-12`",
        desc=("make_totals.py derives data/loans.csv with one parent step "
              "too many, so the primary derivation fails to resolve and the "
              "script reads the stale root-level loans.csv copy instead"),
        why=("the stale annual total is what the shipped artifact and the "
             "paper report")),
}

# A benign confirmed row that matches NO plant signature: it keeps the
# row-deletion sabotage register non-empty, serves as the unrelated
# tombstone target, and is the row the loosened Test-3 signature matches.
BENIGN_ROW = rb.error_row(
    "E-0900", etype="stale_or_wrong_path",
    source="`README.md`", location="`README.md:20`",
    desc="the package README's script order omits the final export step",
    why=("a replicator following the README reproduces an incomplete "
         "artifact set"))


def score(plant_dir, audit_dir):
    """Run the scorer through the production CLI (subprocess)."""
    return rb.run_script(SCORER, "--plant", plant_dir,
                         "--audit-dir", audit_dir)


def rewrite_register(audit, rows):
    """Hand-edit the run's final code-error register file."""
    (audit / "code_error_register.md").write_text(
        rb.register_text("Code-error register", rb.ERROR_COLS, rows),
        encoding="utf-8")


def plant_row(plant, **overrides):
    """The plant's row with Status/Severity overridden by column name."""
    cols = dict(zip(rb.ERROR_COLS, PLANT_ROWS[plant]))
    cols.update(overrides)
    return [cols[c] for c in rb.ERROR_COLS]


@pytest.fixture(scope="module")
def base_runs(tmp_path_factory):
    """One completed synthetic plant run per plant, built by the real
    exporter (make_b9), with the plant row + benign row in the register."""
    runs = {}
    for plant in PLANTS:
        root = tmp_path_factory.mktemp(f"base_{plant}")
        a = rb.make_b9(root, claims_rows=[rb.claims_row("C-0101")],
                       error_rows=[PLANT_ROWS[plant], BENIGN_ROW])
        runs[plant] = a.audit
    return runs


@pytest.fixture
def plant_run(base_runs, tmp_path):
    def make(plant):
        dest = tmp_path / f"audit_{plant}"
        shutil.copytree(base_runs[plant], dest, symlinks=True)
        return dest
    return make


def doctored_plant_dir(tmp_path, plant, mutate):
    """A plant dir whose answer key is a schema-valid weakening."""
    d = tmp_path / "doctored_plant"
    d.mkdir()
    key = json.loads(
        (CHAIN_PLANTS / plant / "expected.json").read_text(encoding="utf-8"))
    mutate(key)
    (d / "expected.json").write_text(json.dumps(key), encoding="utf-8")
    return d


# ------------------------------------------------------- Test 2: good input


@pytest.mark.parametrize("plant", PLANTS)
def test_good_run_scores_green(plant_run, plant):
    audit = plant_run(plant)
    res = score(CHAIN_PLANTS / plant, audit)
    assert res.returncode == 0, res.stdout + res.stderr
    assert f"PLANT GREEN — {plant}" in res.stdout


@pytest.mark.parametrize("plant", PLANTS)
def test_key_effort_map_lockstep(plant):
    """The key's literal equals DEFAULT_EFFORT_MAP by value (drift guard)."""
    key = json.loads(
        (CHAIN_PLANTS / plant / "expected.json").read_text(encoding="utf-8"))
    dispatch = rb.load_script("dispatch_tracking")
    assert key["effort_map"] == dispatch.DEFAULT_EFFORT_MAP


# -------------------------- sabotage drills through the production CLI
# (one per gate leg; register drills hand-edit the register file)


def test_drill_status_flip(plant_run):
    audit = plant_run("comment_closure")
    rewrite_register(audit, [plant_row("comment_closure", Status="not_error"),
                             BENIGN_ROW])
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "row present at" in res.stdout


def test_drill_severity_drop(plant_run):
    audit = plant_run("comment_closure")
    rewrite_register(audit, [plant_row("comment_closure", Severity="1"),
                             BENIGN_ROW])
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "row present at" in res.stdout


def test_drill_row_deletion(plant_run):
    audit = plant_run("comment_closure")
    rewrite_register(audit, [BENIGN_ROW])
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "no register row matches the plant signature" in res.stdout


def test_drill_workbook_stray_sheet(plant_run):
    from openpyxl import load_workbook
    audit = plant_run("comment_closure")
    wb = load_workbook(audit / "code_review.xlsx")
    wb.create_sheet("Notes")
    wb.save(audit / "code_review.xlsx")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "sheet set" in res.stdout


def test_drill_effort_value(plant_run):
    audit = plant_run("comment_closure")
    manifest_path = audit / "_run" / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["effort_map"]["codemap"] = "high"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "effort_map" in res.stdout


# ------------------------------------------- Test 1: remaining bad inputs


def test_hedge_status_is_not_a_pass(plant_run):
    audit = plant_run("comment_closure")
    rewrite_register(
        audit,
        [plant_row("comment_closure", Status="confirmation_needed"),
         BENIGN_ROW])
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "row present at" in res.stdout


def test_tombstone_to_unrelated_row_fails(plant_run):
    """duplicate_of: a confirmed row that does NOT match the signature."""
    audit = plant_run("comment_closure")
    rewrite_register(
        audit,
        [plant_row("comment_closure", Status="duplicate_of:E-0900"),
         BENIGN_ROW])
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "row present at" in res.stdout


def test_tombstone_to_equivalent_target_passes(plant_run):
    """duplicate_of: a target that matches the signature at confirmed
    severity >= 2 scores GREEN — dedup must not fail a passing plant."""
    audit = plant_run("comment_closure")
    equivalent = plant_row("comment_closure")
    equivalent[rb.ERROR_COLS.index("Error ID")] = "E-0210"
    rewrite_register(
        audit,
        [plant_row("comment_closure", Status="duplicate_of:E-0210"),
         equivalent, BENIGN_ROW])
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 0, res.stdout + res.stderr


def test_hidden_column_reinserted(plant_run):
    from openpyxl import load_workbook
    audit = plant_run("comment_closure")
    wb = load_workbook(audit / "code_review.xlsx")
    ws = wb["Paper Claims"]
    ws.insert_cols(4)
    ws.cell(row=1, column=4, value="Used in Text")
    wb.save(audit / "code_review.xlsx")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "hidden column" in res.stdout


def test_code_errors_header_mutated(plant_run):
    from openpyxl import load_workbook
    audit = plant_run("comment_closure")
    wb = load_workbook(audit / "code_review.xlsx")
    wb["Code Errors"].cell(row=1, column=1, value="Error Identifier")
    wb.save(audit / "code_review.xlsx")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "Code Errors header" in res.stdout


@pytest.mark.parametrize("victim,reason", [
    ("code_error_register.md", "code_error_register.md missing"),
    ("code_review.xlsx", "code_review.xlsx missing"),
    ("_run/manifest.json", "manifest.json missing"),
])
def test_missing_input_is_red(plant_run, victim, reason):
    audit = plant_run("comment_closure")
    (audit / victim).unlink()
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert reason in res.stdout


def test_malformed_register_is_red(plant_run):
    audit = plant_run("comment_closure")
    (audit / "code_error_register.md").write_text(
        "# Code-error register\n\ntruncated mid-write, no table\n",
        encoding="utf-8")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "no code-error table parsed" in res.stdout


def test_malformed_workbook_is_red(plant_run):
    audit = plant_run("comment_closure")
    (audit / "code_review.xlsx").write_bytes(b"not a zip archive")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "code_review.xlsx unreadable" in res.stdout


def test_malformed_manifest_is_red(plant_run):
    audit = plant_run("comment_closure")
    (audit / "_run" / "manifest.json").write_text("{truncated",
                                                  encoding="utf-8")
    res = score(CHAIN_PLANTS / "comment_closure", audit)
    assert res.returncode == 1, res.stdout + res.stderr
    assert "manifest.json invalid JSON" in res.stdout


# ------------------- Test 1 (answer-key schema): missing block is exit 2


@pytest.mark.parametrize("field", [
    "class", "run8_id", "signature", "statuses", "min_severity",
    "workbook", "effort_map",
])
def test_key_with_deleted_block_is_exit_2(plant_run, tmp_path, field):
    audit = plant_run("comment_closure")
    doctored = doctored_plant_dir(tmp_path, "comment_closure",
                                  lambda key: key.pop(field))
    res = score(doctored, audit)
    assert res.returncode == 2, res.stdout + res.stderr
    assert field in res.stderr


def test_key_with_deleted_workbook_subfield_is_exit_2(plant_run, tmp_path):
    audit = plant_run("comment_closure")
    doctored = doctored_plant_dir(
        tmp_path, "comment_closure",
        lambda key: key["workbook"].pop("paper_claims_visible"))
    res = score(doctored, audit)
    assert res.returncode == 2, res.stdout + res.stderr
    assert "paper_claims_visible" in res.stderr


def test_key_invalid_json_is_exit_2(plant_run, tmp_path):
    audit = plant_run("comment_closure")
    d = tmp_path / "broken_plant"
    d.mkdir()
    (d / "expected.json").write_text("{truncated", encoding="utf-8")
    res = score(d, audit)
    assert res.returncode == 2, res.stdout + res.stderr


def test_key_missing_is_exit_2(plant_run, tmp_path):
    audit = plant_run("comment_closure")
    d = tmp_path / "empty_plant"
    d.mkdir()
    res = score(d, audit)
    assert res.returncode == 2, res.stdout + res.stderr


# ------------- Test 3: each leg is load-bearing (schema-valid weakenings)


def test_weakened_statuses_turn_status_flip_green(plant_run, tmp_path):
    audit = plant_run("comment_closure")
    rewrite_register(audit, [plant_row("comment_closure", Status="not_error"),
                             BENIGN_ROW])
    doctored = doctored_plant_dir(
        tmp_path, "comment_closure",
        lambda key: key.update(statuses=["confirmed", "not_error"]))
    res = score(doctored, audit)
    assert res.returncode == 0, res.stdout + res.stderr


def test_zero_severity_floor_turns_severity_drop_green(plant_run, tmp_path):
    audit = plant_run("comment_closure")
    rewrite_register(audit, [plant_row("comment_closure", Severity="1"),
                             BENIGN_ROW])
    doctored = doctored_plant_dir(tmp_path, "comment_closure",
                                  lambda key: key.update(min_severity=0))
    res = score(doctored, audit)
    assert res.returncode == 0, res.stdout + res.stderr


def test_loosened_signature_turns_row_deletion_green(plant_run, tmp_path):
    """The presence leg's Test 3: with a signature loosened to match the
    benign row that is still present, deleting the plant row scores GREEN."""
    audit = plant_run("comment_closure")
    rewrite_register(audit, [BENIGN_ROW])
    doctored = doctored_plant_dir(
        tmp_path, "comment_closure",
        lambda key: key.update(signature=[["readme"]]))
    res = score(doctored, audit)
    assert res.returncode == 0, res.stdout + res.stderr


def test_expected_stray_sheet_turns_workbook_sabotage_green(plant_run,
                                                            tmp_path):
    from openpyxl import load_workbook
    audit = plant_run("comment_closure")
    wb = load_workbook(audit / "code_review.xlsx")
    wb.create_sheet("Notes")
    wb.save(audit / "code_review.xlsx")
    doctored = doctored_plant_dir(
        tmp_path, "comment_closure",
        lambda key: key["workbook"].update(
            sheets=key["workbook"]["sheets"] + ["Notes"]))
    res = score(doctored, audit)
    assert res.returncode == 0, res.stdout + res.stderr


def test_matching_weakened_effort_turns_effort_sabotage_green(plant_run,
                                                              tmp_path):
    audit = plant_run("comment_closure")
    manifest_path = audit / "_run" / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["effort_map"]["codemap"] = "high"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    doctored = doctored_plant_dir(
        tmp_path, "comment_closure",
        lambda key: key["effort_map"].update(codemap="high"))
    res = score(doctored, audit)
    assert res.returncode == 0, res.stdout + res.stderr
